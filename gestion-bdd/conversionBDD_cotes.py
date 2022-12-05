import openpyxl
import json


# Importation de la BDD json
tableau = open('bdd_matchs1.json', 'r')
bdd = json.load(tableau)
print("Tableau lu")


# Importation excel
excel = "gestion-bdd/mlb odds 2022.xlsx"
fichier_excel = openpyxl.load_workbook(excel)
tableau = fichier_excel.active
print("Tableau lu.")

liste_equipes_elo = ["ANA", "ARI", "ATL", "BAL",
                     "BOS", "CHC", "CHW", "CIN",
                     "CLE", "COL", "DET", "FLA",
                     "HOU", "KCR", "LAD", "MIL",
                     "MIN", "NYM", "NYY", "OAK",
                     "PHI", "PIT", "SDP", "SEA",
                     "SFG", "STL", "TBD", "TEX",
                     "TOR", "WSN"]

liste_equipes_cotes = ["LAA", "ARI", "ATL", "BAL",
                       "BOS", "CUB", "CWS", "CIN",
                       "CLE", "COL", "DET", "MIA",
                       "HOU", "KAN", "LAD", "MIL",
                       "MIN", "NYM", "NYY", "OAK",
                       "PHI", "PIT", "SDG", "SEA",
                       "SFO", "STL", "TAM", "TEX",
                       "TOR", "WAS"]

conversion_dict = {"LAA": "ANA", "ARI": "ARI", "ATL": "ATL", "BAL": "BAL",
                   "BOS": "BOS", "CUB": "CHC", "CWS": "CHW", "CIN": "CIN",
                   "CLE": "CLE", "COL": "COL", "DET": "DET", "MIA": "FLA",
                   "HOU": "HOU", "KAN": "KCR", "LOS": "LAD", "MIL": "MIL",
                   "MIN": "MIN", "NYM": "NYM", "NYY": "NYY", "OAK": "OAK",
                   "PHI": "PHI", "PIT": "PIT", "SDG": "SDP", "SEA": "SEA",
                   "SFO": "SFG", "STL": "STL", "TAM": "TBD", "TEX": "TEX",
                   "TOR": "TOR", "WAS": "WSN"
                   }


def donne_num_equipe(tag_equipe):
    """Donne le numéro d'une équipe à partir de son tag."""

    return liste_equipes_cotes.index(tag_equipe)


# Lecture du tableau
def recupere_date_tableur(num_ligne, tableau):
    """Récupère la date d'un match du tableau."""

    date = tableau.cell(row=2 * num_ligne, column=1).value
    annee = 2022
    mois = date // 100
    jour = date % 100
    date_formatee = f"{jour}/{mois}/{annee}"
    return date_formatee


def recupere_equipes_tableur(num_ligne, tableau):
    """Récupère les deux équipes d'un match du tableau."""

    equipes = (str(tableau.cell(row=2 * num_ligne, column=4).value),
               str(tableau.cell(row=2 * num_ligne + 1, column=4).value))
    return equipes


def recupere_scores_tableur(num_ligne, tableau):
    """Récupère les scores d'un match du tableau."""

    scores = (tableau.cell(row=2 * num_ligne, column=15).value,
              tableau.cell(row=2 * num_ligne + 1, column=15).value)
    return scores


def recupere_resultat_tableur(num_ligne, tableau):
    """Récupère le résultat d'un match du tableau."""

    score1 = tableau.cell(row=2 * num_ligne, column=15).value
    score2 = tableau.cell(row=2 * num_ligne + 1, column=15).value
    if score1 > score2:
        return 1
    return 0


def recupere_cotes_tableur(num_ligne, tableau):
    """Récupère les côtes d'un match du tableau."""

    cotes = (str(tableau.cell(row=2 * num_ligne, column=16).value),
             str(tableau.cell(row=2 * num_ligne + 1, column=16).value),
             str(tableau.cell(row=2 * num_ligne, column=17).value),
             str(tableau.cell(row=2 * num_ligne + 1, column=17).value))
    return cotes


# Construction json
print((tableau.max_row + 1) / 2)
for num_ligne in range(2, (4474) // 2):
    date = recupere_date_tableur(num_ligne, tableau)
    tag_equipe1, tag_equipe2 = recupere_equipes_tableur(num_ligne, tableau)
    equipe1 = donne_num_equipe(tag_equipe1)
    equipe2 = donne_num_equipe(tag_equipe2)
    score_equipe1, score_equipe2 = recupere_scores_tableur(num_ligne, tableau)
    res_match = recupere_resultat_tableur(num_ligne, tableau)
    cotes = recupere_cotes_tableur(num_ligne, tableau)
    cote_open1, cote_open2, cote_close1, cote_close2 = cotes
    cle_dict = f"{date}/{equipe2}"
    if cle_dict in bdd:
        bdd.update(
            {cle_dict: {
                "equipe1": equipe2,
                "equipe2": equipe1,
                "score1": score_equipe2,
                "score2": score_equipe1,
                "resultat": 1 - res_match,
                "coteOpen1": cote_open2,
                "coteOpen2": cote_open1,
                "coteClose1": cote_close2,
                "coteClose2": cote_close1}})
    else:
        cle_dict = f"{date}/{equipe1}"
        bdd.update(
            {cle_dict: {
                "equipe1": equipe1,
                "equipe2": equipe2,
                "score1": score_equipe1,
                "score2": score_equipe2,
                "resultat": res_match,
                "coteOpen1": cote_open1,
                "coteOpen2": cote_open2,
                "coteClose1": cote_close1,
                "coteClose2": cote_close2}})


with open("bdd_matchs2.json", "w", encoding="utf-8") as bdd_json:
    json.dump(bdd, bdd_json, indent=4, default=str)
