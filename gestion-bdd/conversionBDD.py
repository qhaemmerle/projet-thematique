import openpyxl
import json
import datetime


# Importation excel
excel = "BD_baseball.xlsx"
fichier_excel = openpyxl.load_workbook(excel)
tableau = fichier_excel.active
print("Tableau lu.")


liste_equipes_elo = ["ANA", "ARI", "ATL", "BAL",
                     "BOS", "CHC", "CHW", "CIN",
                     "CLE", "COL", "DET", "FLA",
                     "HOU", "KCR", "LAD", "MIL",
                     "MIN", "NYM", "NYY", "OAK",
                     "PHI", "PIT", "SDP", "SEA",
                     "SFG", "STL", "TBD", "TEX",
                     "TOR", "WSN"]


def donne_num_equipe(tag_equipe):
    """Donne le numéro d'une équipe à partir de son tag."""

    return liste_equipes_elo.index(tag_equipe)


# Lecture du tableau
def recupere_date_tableur(num_ligne, tableau):
    """Récupère la date d'un match du tableau."""

    date = tableau.cell(row=num_ligne, column=1).value
    if type(date) != datetime.datetime:
        return ""
    annee = date.year
    mois = date.month
    jour = date.day
    date_formatee = f"{jour}/{mois}/{annee}"
    return date_formatee


def recupere_equipes_tableur(num_ligne, tableau):
    """Récupère les deux équipes d'un match du tableau."""

    equipes = (str(tableau.cell(row=num_ligne, column=5).value),
               str(tableau.cell(row=num_ligne, column=6).value))
    return equipes


def recupere_scores_tableur(num_ligne, tableau):
    scores = (tableau.cell(row=num_ligne, column=25).value,
              tableau.cell(row=num_ligne, column=26).value)
    return scores


def recupere_resultat_tableur(num_ligne, tableau):
    """Récupère le résultat d'un match du tableau."""

    score1 = tableau.cell(row=num_ligne, column=25).value
    score2 = tableau.cell(row=num_ligne, column=26).value
    if score1 > score2:
        return 1
    return 0


# Construction json
dictionary = {}
for num_ligne in range(2, tableau.max_row + 1):
    date = recupere_date_tableur(num_ligne, tableau)
    tag_equipe1, tag_equipe2 = recupere_equipes_tableur(num_ligne, tableau)
    equipe1 = donne_num_equipe(tag_equipe1)
    equipe2 = donne_num_equipe(tag_equipe2)
    score_equipe1, score_equipe2 = recupere_scores_tableur(num_ligne, tableau)
    res_match = recupere_resultat_tableur(num_ligne, tableau)
    cle_dict = f"{date}/{equipe1}"
    dictionary.update(
        {cle_dict: {
            "equipe1": equipe1,
            "equipe2": equipe2,
            "score1": score_equipe1,
            "score2": score_equipe2,
            "resultat": res_match}})


with open("exemple.json", "w", encoding="utf-8") as bdd_json:
    json.dump(dictionary, bdd_json, indent=4, default=str)

print(type(recupere_date_tableur(3, tableau)))
